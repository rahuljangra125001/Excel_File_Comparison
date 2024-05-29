import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox


def column_to_set(file_path, column_name):

    # Read the Excel file into a DataFrame
    try:
        df = pd.read_excel(file_path, usecols=column_name)
    except Exception as e:
        raise ValueError(f"Error reading file '{file_path}': {e}")

    return df.values.flatten()


def compare_and_color(file1_path, file2_path, column_to_compare):
    # Load workbooks and select the first sheet
    wb1 = load_workbook(file1_path)
    ws1 = wb1.active

    wb2 = load_workbook(file2_path)
    ws2 = wb2.active

    # Define fill colors
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    # Function to apply fill color to cells based on presence in both sets
    def apply_fill(ws, other_set, fill_yellow, fill_red, column_letter):
        for row in range(2, ws.max_row + 1):  # Start from 2 to skip the header
            cell = ws[f"{column_letter}{row}"]
            value = cell.value
            if value in other_set:
                cell.fill = fill_yellow
            else:
                cell.fill = fill_red
    for i in column_to_compare:
        # Convert columns to sets for comparison
        set1 = column_to_set(file1_path, i)
        set2 = column_to_set(file2_path, i)
        # Apply fill to cells in file1.xlsx based on comparison with file2.xlsx
        apply_fill(ws1, set2, yellow_fill, red_fill, i)

        # Apply fill to cells in file2.xlsx based on comparison with file1.xlsx
        apply_fill(ws2, set1, yellow_fill, red_fill, i)

    # Save the workbooks with the new formatting
    wb1.save('file1_comparison.xlsx')
    wb2.save('file2_comparison.xlsx')


def browse_file(entry):
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if filename:
        entry.delete(0, tk.END)
        entry.insert(0, filename)


def run_comparison():
    file1_path = file1_entry.get()
    file2_path = file2_entry.get()
    columns = columns_entry.get().split(",")
    columns = [col.strip() for col in columns]

    if not file1_path or not file2_path or not columns:
        messagebox.showerror("Input Error", "All fields are required.")
        return

    try:
        compare_and_color(file1_path, file2_path, columns)
        messagebox.showinfo("Success", "Comparison completed and files saved.")
    except ValueError as ve:
        messagebox.showerror("Validation Error", str(ve))
    except FileNotFoundError as fe:
        messagebox.showerror("File Error", str(fe))
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")


# Create the GUI
root = tk.Tk()
root.title("Excel Column Comparison")

tk.Label(root, text="File 1 Path:").grid(row=0, column=0, padx=10, pady=5, sticky="e")
file1_entry = tk.Entry(root, width=50)
file1_entry.grid(row=0, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: browse_file(file1_entry)).grid(row=0, column=2, padx=10, pady=5)

tk.Label(root, text="File 2 Path:").grid(row=1, column=0, padx=10, pady=5, sticky="e")
file2_entry = tk.Entry(root, width=50)
file2_entry.grid(row=1, column=1, padx=10, pady=5)
tk.Button(root, text="Browse", command=lambda: browse_file(file2_entry)).grid(row=1, column=2, padx=10, pady=5)

tk.Label(root, text="Columns to Compare (comma-separated):").grid(row=4, column=0, padx=10, pady=5, sticky="e")
columns_entry = tk.Entry(root, width=50)
columns_entry.grid(row=4, column=1, padx=10, pady=5)

tk.Button(root, text="Compare", command=run_comparison).grid(row=5, column=0, columnspan=3, pady=10)

root.mainloop()
